import requests

from openpyxl import load_workbook

from . import models

BASE_URL = "http://45.9.27.241:8000/"


def upload_file(file):
    url = "http://45.9.27.241:8000/file_upload/"
    user_file = {
        "file": file
    }
    response = requests.post(url, json=user_file)
    print(response)


def get_result_file(file, user):
    url = f"http://45.9.27.241:8000/process_xls/{user}/"


def import_spgz_data_from_excel(file):
    """Импорт данных"""
    book = load_workbook(file)
    sheet = book.active

    for row in range(2, sheet.max_row):
        # Читаем значения из excel файла
        ID = sheet[row][0].value
        kpzg = sheet[row][3].value
        spgz_name = sheet[row][4].value
        per_unit = sheet[row][6].value
        okpd = sheet[row][7].value
        okpd2 = sheet[row][8].value

        # Сохраняем в БД
        models.SPGZ.objects.update_or_create(
            id=ID, kpzg=kpzg, spgz_name=spgz_name, unit=per_unit, okpd=okpd, okpd2=okpd2
        )


def import_tcn_data_from_excel(file):
    """Импорт данных из excel файла"""
    book = load_workbook(file)
    sheet = book.active

    for row in range(2, sheet.max_row):
        cipher = sheet[row][0].value
        name = sheet[row][0].value
        unit = sheet[row][0].value
        quantity = sheet[row][0].value
        spgz_name = sheet[row][0].value
        id = sheet[row][0].value
        kpzg = sheet[row][0].value
        unit2 = sheet[row][0].value
        quantity2 = sheet[row][0].value

    models.TSN.objects.update_or_create(
        cipher=cipher, name=name, unit=unit, quantity=quantity, spgz_name=spgz_name, id=id, kpzg=kpzg, unit2=unit2,
        quantity2=quantity2,
    )


def import_sn_data_from_excel(file):
    """Импорт данных из excel файла"""
    book = load_workbook(file)
    sheet = book.active

    for row in range(2, sheet.max_row):
        cipher = sheet[row][0].value
        name = sheet[row][0].value
        unit = sheet[row][0].value
        quantity = sheet[row][0].value
        spgz_name = sheet[row][0].value
        id = sheet[row][0].value
        kpzg = sheet[row][0].value
        unit2 = sheet[row][0].value
        quantity2 = sheet[row][0].value

    models.SN.objects.update_or_create(
        cipher=cipher, name=name, unit=unit, quantity=quantity, spgz_name=spgz_name, id=id, kpzg=kpzg, unit2=unit2,
        quantity2=quantity2,
    )


def import_template_data_from_excel(file):
    book = load_workbook(file)
    sheet = book.active
    for row in range(2, sheet.max_row):
        pass

    pass
