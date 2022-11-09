from config.celery import app
from openpyxl import load_workbook

from . import models


@app.task
def import_spgz_data_from_excel(file):
    """Импорт данных"""
    files = models.SPGZFile.objects.all()

    book = load_workbook(file)
    sheet = book.active

    for row in range(2, sheet.max_row):
        # Читаем значения из excel файла
        ID = sheet[row][0].value
        kpzg = sheet[row][3].value
        spgz_name = sheet[row][4].value
        per_unit = sheet[row][6].value
        okpd = sheet[row][7].value
        okpd2 = sheet[row][8].value
        print(ID, spgz_name)

        # Сохраняем в БД
        models.SPGZ.objects.update_or_create(
            id=ID, kpzg=kpzg, spgz_name=spgz_name, unit=per_unit, okpd=okpd, okpd2=okpd2
        )


@app.task
def import_tsn_data_from_excel(file):
    pass


@app.task
def import_sn_data_from_excel(file):
    pass


@app.task
def import_template_data_from_excel(file):
    pass
